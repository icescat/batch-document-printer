"""
页数统计对话框
显示文档页数统计结果和导出功能
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path
from typing import List
import csv
import json
from datetime import datetime
import threading

from src.core.page_count_manager import PageCountManager, PageCountSummary, PageCountResult, PageCountStatus
from src.core.models import Document, FileType


class PageCountProgressDialog:
    """页数计算进度对话框"""
    
    def __init__(self, parent):
        """初始化进度对话框"""
        self.parent = parent
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("页数计算中，遇到密码保护文档必须手动关掉打开的文档")
        self.dialog.geometry("400x180")
        self.dialog.resizable(False, False)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # 居中显示
        self._center_dialog()
        
        # 创建界面
        self._create_widgets()
        
        # 取消标志
        self.cancelled = False
    
    def _center_dialog(self):
        """将对话框居中显示"""
        self.dialog.update_idletasks()
        
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        dialog_width = self.dialog.winfo_reqwidth()
        dialog_height = self.dialog.winfo_reqheight()
        
        x = parent_x + (parent_width - dialog_width) // 2
        y = parent_y + (parent_height - dialog_height) // 2
        
        self.dialog.geometry(f"+{x}+{y}")
    
    def _create_widgets(self):
        """创建界面组件"""
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill="both", expand=True)
        
        # 状态标签
        self.status_label = ttk.Label(main_frame, text="正在准备计算...")
        self.status_label.pack(pady=(0, 10))
        
        # 进度文本（放在进度条上方）
        self.progress_label = ttk.Label(main_frame, text="0/0", font=("", 9))
        self.progress_label.pack(pady=(0, 5))
        
        # 进度条
        self.progress_bar = ttk.Progressbar(
            main_frame, mode="determinate", length=300
        )
        self.progress_bar.pack(pady=(0, 15))
        
        # 取消按钮
        self.cancel_button = ttk.Button(
            main_frame, text="取消", command=self._on_cancel
        )
        self.cancel_button.pack()
    
    def update_progress(self, current: int, total: int, message: str):
        """更新进度"""
        if self.dialog.winfo_exists():
            progress = (current / total) * 100 if total > 0 else 0
            self.progress_bar['value'] = progress
            self.progress_label.config(text=f"{current}/{total}")
            self.status_label.config(text=message)
            self.dialog.update()
    
    def _on_cancel(self):
        """取消计算"""
        self.cancelled = True
        self.close()
    
    def close(self):
        """关闭对话框"""
        if self.dialog.winfo_exists():
            self.dialog.destroy()


class PageCountResultDialog:
    """页数统计结果对话框"""
    
    def __init__(self, parent, summary: PageCountSummary):
        """
        初始化结果对话框
        
        Args:
            parent: 父窗口
            summary: 页数统计汇总
        """
        self.parent = parent
        self.summary = summary
        
        # 创建对话框窗口
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("页数统计结果")
        self.dialog.geometry("750x450")
        self.dialog.resizable(True, True)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # 居中显示
        self._center_dialog()
        
        # 创建界面
        self._create_widgets()
        
        # 等待对话框关闭
        self.dialog.wait_window()
    
    def _center_dialog(self):
        """将对话框居中显示"""
        self.dialog.update_idletasks()
        
        parent_x = self.parent.winfo_rootx()
        parent_y = self.parent.winfo_rooty()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        
        dialog_width = self.dialog.winfo_reqwidth()
        dialog_height = self.dialog.winfo_reqheight()
        
        x = parent_x + (parent_width - dialog_width) // 2
        y = parent_y + (parent_height - dialog_height) // 2
        
        self.dialog.geometry(f"+{x}+{y}")
    
    def _create_widgets(self):
        """创建界面组件"""
        # 主框架
        main_frame = ttk.Frame(self.dialog, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # 统计概览区域
        self._create_overview_section(main_frame)
        
        # 文件类型统计区域
        self._create_file_type_section(main_frame)
        
        # 计算问题区域
        self._create_problems_section(main_frame)
        
        # 按钮区域（独立的框架，固定在底部）
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", side="bottom", pady=(10, 0))
        self._create_button_section(button_frame)
        
        # 设置权重
        main_frame.grid_rowconfigure(1, weight=1)  # 问题区域可扩展
    
    def _create_overview_section(self, parent):
        """创建统计概览区域"""
        # 统计概览框架
        overview_frame = ttk.LabelFrame(parent, text="📊 统计概览", padding="10")
        overview_frame.pack(fill="x", pady=(0, 10))
        
        # 总体统计（横向排列）
        col = 0
        
        # 总文档数
        ttk.Label(overview_frame, text="总文档数:", font=("", 10, "bold")).grid(
            row=0, column=col, sticky="w", padx=(0, 5))
        ttk.Label(overview_frame, text=f"{self.summary.total_files} 个", 
                 font=("", 10)).grid(row=0, column=col+1, sticky="w", padx=(0, 20))
        col += 2
        
        # 总页数
        ttk.Label(overview_frame, text="总页数:", font=("", 10, "bold")).grid(
            row=0, column=col, sticky="w", padx=(0, 5))
        
        # 检查是否有Excel文件参与统计，如果有则添加"约"字
        has_excel = self.summary.excel_files > 0
        total_pages_text = f"约{self.summary.total_pages:,} 页" if has_excel else f"{self.summary.total_pages:,} 页"
        
        ttk.Label(overview_frame, text=total_pages_text, 
                 font=("", 10)).grid(row=0, column=col+1, sticky="w", padx=(0, 20))
        col += 2
        
        # 成功数
        ttk.Label(overview_frame, text="成功:", font=("", 10, "bold"), 
                 foreground="green").grid(row=0, column=col, sticky="w", padx=(0, 5))
        ttk.Label(overview_frame, text=f"{self.summary.success_count} 个", 
                 font=("", 10)).grid(row=0, column=col+1, sticky="w", padx=(0, 20))
        col += 2
        
        # 跳过数（始终显示）
        ttk.Label(overview_frame, text="跳过:", font=("", 10, "bold"), 
                 foreground="orange").grid(row=0, column=col, sticky="w", padx=(0, 5))
        ttk.Label(overview_frame, text=f"{self.summary.skipped_count} 个", 
                 font=("", 10)).grid(row=0, column=col+1, sticky="w", padx=(0, 20))
        col += 2
        
        # 失败数（始终显示）
        ttk.Label(overview_frame, text="失败:", font=("", 10, "bold"), 
                 foreground="red").grid(row=0, column=col, sticky="w", padx=(0, 5))
        ttk.Label(overview_frame, text=f"{self.summary.error_count} 个", 
                 font=("", 10)).grid(row=0, column=col+1, sticky="w")
        
    
    def _create_file_type_section(self, parent):
        """创建文件类型统计区域"""
        # 文件类型统计框架
        type_frame = ttk.LabelFrame(parent, text="📄 按文件类型统计", padding="10")
        type_frame.pack(fill="x", pady=(0, 10))
        
        # 横向排列所有文件类型统计
        col = 0
        
        # Word统计
        ttk.Label(type_frame, text="Word:", font=("", 10, "bold")).grid(
            row=0, column=col, sticky="w", padx=(0, 5))
        ttk.Label(type_frame, 
                 text=f"{self.summary.word_files}个 / {self.summary.word_pages:,}页",
                 font=("", 10)).grid(row=0, column=col+1, sticky="w", padx=(0, 20))
        col += 2
        
        # PPT统计
        ttk.Label(type_frame, text="PPT:", font=("", 10, "bold")).grid(
            row=0, column=col, sticky="w", padx=(0, 5))
        ttk.Label(type_frame, 
                 text=f"{self.summary.ppt_files}个 / {self.summary.ppt_pages:,}页",
                 font=("", 10)).grid(row=0, column=col+1, sticky="w", padx=(0, 20))
        col += 2
        
        # Excel统计
        ttk.Label(type_frame, text="Excel:", font=("", 10, "bold")).grid(
            row=0, column=col, sticky="w", padx=(0, 5))
        
        # Excel页数显示时添加"约"字，提醒用户不确定性
        excel_pages_text = f"{self.summary.excel_files}个 / 约{self.summary.excel_pages:,}页" if self.summary.excel_files > 0 else f"{self.summary.excel_files}个 / {self.summary.excel_pages:,}页"
        
        ttk.Label(type_frame, 
                 text=excel_pages_text,
                 font=("", 10)).grid(row=0, column=col+1, sticky="w", padx=(0, 20))
        col += 2
        
        # PDF统计
        ttk.Label(type_frame, text="PDF:", font=("", 10, "bold")).grid(
            row=0, column=col, sticky="w", padx=(0, 5))
        ttk.Label(type_frame, 
                 text=f"{self.summary.pdf_files}个 / {self.summary.pdf_pages:,}页",
                 font=("", 10)).grid(row=0, column=col+1, sticky="w", padx=(0, 20))
        col += 2
        
        # 图片统计
        ttk.Label(type_frame, text="图片:", font=("", 10, "bold")).grid(
            row=0, column=col, sticky="w", padx=(0, 5))
        ttk.Label(type_frame, 
                 text=f"{self.summary.image_files}个 / {self.summary.image_pages:,}页",
                 font=("", 10)).grid(row=0, column=col+1, sticky="w")
    
    def _create_problems_section(self, parent):
        """创建计算问题区域"""
        
        # 问题文件框架
        problems_frame = ttk.LabelFrame(parent, text="⚠️ 问题文件详情", padding="10")
        problems_frame.pack(fill="both", expand=True, pady=(0, 10))
        
        # 问题文件列表
        problem_files = self.summary.skipped_files + self.summary.error_files
        
        # 创建树形视图容器
        tree_frame = ttk.Frame(problems_frame)
        tree_frame.pack(fill="both", expand=True)
        
        # 创建树形视图（简化版本，只显示基本信息）
        columns = ("文件名", "文件类型", "文件路径")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=5)
        
        # 设置列标题和宽度
        tree.heading("文件名", text="文件名")
        tree.heading("文件类型", text="文件类型")
        tree.heading("文件路径", text="文件路径")
        
        tree.column("文件名", width=250)
        tree.column("文件类型", width=100)
        tree.column("文件路径", width=350)
        
        if problem_files:
            # 添加问题文件
            for result in problem_files:
                file_type_display = {
                    FileType.WORD: "Word文档",
                    FileType.PPT: "PowerPoint",
                    FileType.EXCEL: "Excel表格",
                    FileType.PDF: "PDF文件",
                    FileType.IMAGE: "图片文件"
                }.get(result.document.file_type, "未知类型")
                
                tree.insert("", "end", values=(
                    result.document.file_name,
                    file_type_display,
                    str(result.document.file_path)
                ))
        else:
            # 没有问题文件时显示提示
            tree.insert("", "end", values=(
                "✅ 所有文件均成功计算页数",
                "—",
                "—"
            ))
        
        # 滚动条
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # 布局（与主界面一致的Grid布局）
        tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # 提示信息
        if problem_files:
            tip_label = ttk.Label(problems_frame, 
                                 text=f"共 {len(problem_files)} 个文件存在问题，可导出详细报告查看完整信息",
                                 font=("", 9), foreground="gray")
            tip_label.pack(anchor="w", pady=(5, 0))
        else:
            tip_label = ttk.Label(problems_frame, 
                                 text="🎉 恭喜！所有文件页数统计均成功完成",
                                 font=("", 9), foreground="green")
            tip_label.pack(anchor="w", pady=(5, 0))
    
    def _get_problem_description(self, result: PageCountResult) -> str:
        """获取问题描述（简化版）"""
        error_msg = result.error_message.lower()
        if ("密码" in error_msg or "password" in error_msg or "保护" in error_msg or 
            "加密" in error_msg or "权限" in error_msg or "access" in error_msg):
            return "🔒 文件被加密或权限受限"
        else:
            return "💥 文件已损坏或格式不正确"
    
    def _create_button_section(self, parent):
        """创建按钮区域"""
        # 直接使用传入的parent作为按钮容器
        
        # 导出全部按钮
        self.btn_export_all = ttk.Button(
            parent, text="导出全部", 
            command=self._export_all_report
        )
        self.btn_export_all.pack(side="left", padx=(0, 10))
        
        # 导出错误报告按钮
        if self.summary.skipped_count > 0 or self.summary.error_count > 0:
            self.btn_export_errors = ttk.Button(
                parent, text="导出错误报告", 
                command=self._export_error_report
            )
            self.btn_export_errors.pack(side="left", padx=(0, 10))
        
        # 关闭按钮
        self.btn_close = ttk.Button(
            parent, text="关闭", 
            command=self._on_close
        )
        self.btn_close.pack(side="right")
    
    def _export_all_report(self):
        """导出完整报告"""
        try:
            file_path = filedialog.asksaveasfilename(
                title="保存完整统计报告",
                defaultextension=".csv",
                filetypes=[
                    ("CSV文件", "*.csv"),
                    ("Excel文件", "*.xlsx"),
                    ("文本文件", "*.txt")
                ]
            )
            
            if not file_path:
                return
            
            file_path = Path(file_path)
            
            if file_path.suffix.lower() == '.csv':
                self._export_to_csv(file_path, include_all=True)
            elif file_path.suffix.lower() == '.xlsx':
                self._export_to_excel(file_path, include_all=True)
            else:
                self._export_to_text(file_path, include_all=True)
            
            # 去掉导出成功提示窗口
            print(f"报告已导出到: {file_path}")
            
        except PermissionError:
            messagebox.showerror("导出失败", "文件正在被其他程序使用，请关闭相关文件后重试")
        except FileNotFoundError:
            messagebox.showerror("导出失败", "指定的文件路径不存在，请选择有效的保存位置")
        except Exception as e:
            error_msg = str(e)
            if "openpyxl" in error_msg.lower():
                messagebox.showerror("导出失败", "缺少Excel支持库，Excel导出功能不可用，请选择CSV或文本格式")
            elif "permission" in error_msg.lower() or "access" in error_msg.lower():
                messagebox.showerror("导出失败", "没有写入权限，请选择其他保存位置或以管理员身份运行")
            elif "disk" in error_msg.lower() or "space" in error_msg.lower():
                messagebox.showerror("导出失败", "磁盘空间不足，请清理磁盘空间后重试")
            else:
                messagebox.showerror("导出失败", f"导出过程中出现错误，请检查文件路径和权限\n\n技术详情：{error_msg}")
    
    def _export_error_report(self):
        """导出错误报告"""
        try:
            file_path = filedialog.asksaveasfilename(
                title="保存错误报告",
                defaultextension=".csv",
                filetypes=[
                    ("CSV文件", "*.csv"),
                    ("文本文件", "*.txt")
                ]
            )
            
            if not file_path:
                return
            
            file_path = Path(file_path)
            
            if file_path.suffix.lower() == '.csv':
                self._export_to_csv(file_path, include_all=False)
            else:
                self._export_to_text(file_path, include_all=False)
            
            # 去掉导出成功提示窗口
            print(f"错误报告已导出到: {file_path}")
            
        except PermissionError:
            messagebox.showerror("导出失败", "文件正在被其他程序使用，请关闭相关文件后重试")
        except FileNotFoundError:
            messagebox.showerror("导出失败", "指定的文件路径不存在，请选择有效的保存位置")
        except Exception as e:
            error_msg = str(e)
            if "permission" in error_msg.lower() or "access" in error_msg.lower():
                messagebox.showerror("导出失败", "没有写入权限，请选择其他保存位置或以管理员身份运行")
            elif "disk" in error_msg.lower() or "space" in error_msg.lower():
                messagebox.showerror("导出失败", "磁盘空间不足，请清理磁盘空间后重试")
            else:
                messagebox.showerror("导出失败", f"导出过程中出现错误，请检查文件路径和权限\n\n技术详情：{error_msg}")
    
    def _export_to_csv(self, file_path: Path, include_all: bool = True):
        """导出到CSV文件"""
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # 写入标题
            if include_all:
                writer.writerow(['文件名', '文件类型', '页数', '状态', '文件路径', '错误信息'])
                
                # 写入所有文件数据
                all_results = []
                
                # 收集所有结果
                for result in self.summary.skipped_files + self.summary.error_files:
                    all_results.append(result)
                
                # 添加成功的文件（需要从原始数据重构，这里简化处理）
                # 由于我们没有保存成功的详细结果，这里只显示问题文件
                for result in all_results:
                    file_type_display = {
                        FileType.WORD: "Word文档",
                        FileType.PPT: "PowerPoint",
                        FileType.EXCEL: "Excel表格",
                        FileType.PDF: "PDF文件",
                        FileType.IMAGE: "图片文件"
                    }.get(result.document.file_type, "未知")
                    
                    status_display = "成功" if result.page_count is not None else self._get_problem_description(result)
                    
                    # 如果是Excel文件且有页数，在页数前添加"约"字
                    page_count_display = result.page_count or "N/A"
                    if result.page_count is not None and result.document.file_type == FileType.EXCEL:
                        page_count_display = f"约{result.page_count}"
                    
                    writer.writerow([
                        result.document.file_name,
                        file_type_display,
                        page_count_display,
                        status_display,
                        str(result.document.file_path),
                        result.error_message
                    ])
            else:
                # 只导出错误文件
                writer.writerow(['文件名', '文件类型', '文件路径'])
                
                for result in self.summary.skipped_files + self.summary.error_files:
                    file_type_display = {
                        FileType.WORD: "Word文档",
                        FileType.PPT: "PowerPoint",
                        FileType.EXCEL: "Excel表格",
                        FileType.PDF: "PDF文件",
                        FileType.IMAGE: "图片文件"
                    }.get(result.document.file_type, "未知类型")
                    
                    writer.writerow([
                        result.document.file_name,
                        file_type_display,
                        str(result.document.file_path)
                    ])
    
    def _export_to_text(self, file_path: Path, include_all: bool = True):
        """导出到文本文件"""
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("页数统计报告\n")
            f.write("=" * 50 + "\n")
            f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            # 统计概览
            f.write("统计概览:\n")
            f.write(f"  总文档数: {self.summary.total_files} 个\n")
            
            # 如果有Excel文件参与统计，在总页数前添加"约"字
            total_pages_text = f"约{self.summary.total_pages:,}" if self.summary.excel_files > 0 else f"{self.summary.total_pages:,}"
            f.write(f"  总页数: {total_pages_text} 页\n")
            
            f.write(f"  成功计算: {self.summary.success_count} 个\n")
            f.write(f"  跳过文件: {self.summary.skipped_count} 个\n")
            f.write(f"  计算失败: {self.summary.error_count} 个\n\n")
            
            # 按类型统计
            if self.summary.word_files > 0 or self.summary.ppt_files > 0 or \
               self.summary.excel_files > 0 or self.summary.pdf_files > 0 or \
               self.summary.image_files > 0:
                f.write("按文件类型统计:\n")
                if self.summary.word_files > 0:
                    f.write(f"  Word文档: {self.summary.word_files}个文件, {self.summary.word_pages:,}页\n")
                if self.summary.ppt_files > 0:
                    f.write(f"  PowerPoint: {self.summary.ppt_files}个文件, {self.summary.ppt_pages:,}页\n")
                if self.summary.excel_files > 0:
                    f.write(f"  Excel表格: {self.summary.excel_files}个文件, 约{self.summary.excel_pages:,}页 (估算)\n")
                if self.summary.pdf_files > 0:
                    f.write(f"  PDF文件: {self.summary.pdf_files}个文件, {self.summary.pdf_pages:,}页\n")
                if self.summary.image_files > 0:
                    f.write(f"  图片文件: {self.summary.image_files}个文件, {self.summary.image_pages:,}页\n")
                f.write("\n")
            
            # 问题文件列表
            if not include_all:
                f.write("问题文件详情:\n")
                f.write("-" * 30 + "\n")
                
                for result in self.summary.skipped_files + self.summary.error_files:
                    f.write(f"文件名: {result.document.file_name}\n")
                    f.write(f"类型: {result.document.file_type.value}\n")
                    f.write(f"路径: {result.document.file_path}\n")
                    f.write("-" * 30 + "\n")
    
    def _export_to_excel(self, file_path: Path, include_all: bool = True):
        """导出到Excel文件（需要openpyxl库）"""
        try:
            import openpyxl  # type: ignore
            from openpyxl.styles import Font, PatternFill  # type: ignore
            
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is not None:  # 类型安全检查
                ws.title = "页数统计报告"  # type: ignore
                
                # 设置标题
                headers = ['文件名', '文件类型', '文件路径']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)  # type: ignore
                    cell.font = Font(bold=True)  # type: ignore
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # type: ignore
                
                # 写入数据（简化版本，只写入问题文件）
                row = 2
                for result in self.summary.skipped_files + self.summary.error_files:
                    file_type_display = {
                        FileType.WORD: "Word文档",
                        FileType.PPT: "PowerPoint",
                        FileType.EXCEL: "Excel表格",
                        FileType.PDF: "PDF文件",
                        FileType.IMAGE: "图片文件"
                    }.get(result.document.file_type, "未知")
                    
                    ws.cell(row=row, column=1, value=result.document.file_name)  # type: ignore
                    ws.cell(row=row, column=2, value=file_type_display)  # type: ignore
                    ws.cell(row=row, column=3, value=str(result.document.file_path))  # type: ignore
                    row += 1
                
                # 调整列宽
                ws.column_dimensions['A'].width = 30  # type: ignore
                ws.column_dimensions['B'].width = 15  # type: ignore
                ws.column_dimensions['C'].width = 50  # type: ignore
            
            wb.save(file_path)
            
        except ImportError:
            # 如果没有openpyxl，回退到CSV
            csv_path = file_path.with_suffix('.csv')
            self._export_to_csv(csv_path, include_all)
            raise Exception(f"未安装openpyxl库，已导出为CSV格式: {csv_path}")
    
    def _on_close(self):
        """关闭对话框"""
        self.dialog.destroy()


def show_page_count_dialog(parent, documents: List[Document]):
    """
    显示页数统计对话框的便捷函数
    
    Args:
        parent: 父窗口
        documents: 文档列表
    """
    if not documents:
        messagebox.showwarning("提示", "没有要统计的文档")
        return
    
    # 创建进度对话框
    progress_dialog = PageCountProgressDialog(parent)
    
    # 创建页数管理器
    page_manager = PageCountManager()
    page_manager.set_progress_callback(progress_dialog.update_progress)
    
    summary = None
    error_message = None
    
    def calculate_in_background():
        """在后台线程中计算页数"""
        nonlocal summary, error_message
        try:
            summary = page_manager.calculate_all_pages(documents)
        except Exception as e:
            error_message = str(e)
        finally:
            # 关闭进度对话框
            if not progress_dialog.cancelled:
                progress_dialog.close()
    
    # 启动后台计算
    thread = threading.Thread(target=calculate_in_background, daemon=True)
    thread.start()
    
    # 等待计算完成或用户取消
    parent.wait_window(progress_dialog.dialog)
    
    if progress_dialog.cancelled:
        page_manager.cancel_calculation()
        return
    
    # 等待线程完成
    thread.join(timeout=1.0)
    
    if error_message:
        messagebox.showerror("错误", f"页数计算失败: {error_message}")
        return
    
    if summary:
        # 显示结果对话框
        PageCountResultDialog(parent, summary) 